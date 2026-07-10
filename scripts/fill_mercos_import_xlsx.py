"""Preenche modelo de importação Mercos com produtos do Supabase."""
import re
import shutil
import sys
from pathlib import Path

import openpyxl

BACKEND = Path(__file__).resolve().parents[1].parent / "backend-xnamai"
sys.path.insert(0, str(BACKEND))

from dotenv import load_dotenv

load_dotenv(BACKEND / ".env")

from app.services.supabase_service import supabase  # noqa: E402

NCM_MAP = [
    (r"lampada|lâmpada", "94052100"),
    (r"poltrona", "94016100"),
    (r"macarr", "19022000"),
    (r"tempero", "44219000"),
    (r"mesa", "94036000"),
    (r"toalha", "63026000"),
    (r"perfume", "33030010"),
    (r"camisa", "62052000"),
    (r"bolsa", "42022100"),
    (r"cama pet", "94049000"),
    (r"trenzinho", "95030099"),
    (r"torneira", "84818019"),
    (r"porcelanato", "69072100"),
    (r"caixa de som|bluetooth", "85182200"),
    (r"cosm|skincare", "33049990"),
    (r"vinho", "22042100"),
]

CATEGORY_MAP = [
    (r"mesa|poltrona", "Móveis"),
    (r"macarr", "Alimentos"),
    (r"tempero", "Utilidades Domésticas"),
    (r"lampada|lâmpada", "Iluminação"),
    (r"toalha", "Têxtil"),
    (r"perfume", "Perfumaria"),
    (r"camisa", "Vestuário"),
    (r"bolsa", "Acessórios"),
    (r"cama pet", "Pet"),
    (r"trenzinho", "Brinquedos"),
    (r"torneira", "Construção"),
    (r"porcelanato", "Revestimentos"),
    (r"caixa de som", "Eletrônicos"),
    (r"cosm|skincare", "Cosméticos"),
    (r"vinho", "Bebidas"),
]


def pick(patterns: list[tuple[str, str]], name: str, default: str = "") -> str:
    for pat, val in patterns:
        if re.search(pat, name, re.I):
            return val
    return default


def make_code(nome: str, idx: int) -> str:
    clean = re.sub(r"\[Exemplo\]\s*", "", nome, flags=re.I).strip()
    words = re.findall(r"[A-Za-z0-9]+", clean)
    parts: list[str] = []
    for word in words[:3]:
        parts.append(word if word.isdigit() else word[:4].upper())
    return ("-".join(parts) if parts else f"PROD-{idx:03d}")[:30]


def main() -> None:
    src = Path(r"c:\Users\vicel\Downloads\modelo_importacao_produto_com_ncm (2).xlsx")
    out = Path(r"c:\Users\vicel\Downloads\produtos_mercos_preenchido.xlsx")
    shutil.copy2(src, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Planilha1"]

    rows = (
        supabase.table("produtos")
        .select("codigo,nome,preco_tabela,preco_minimo,saldo_estoque,unidade,descricao,ativo")
        .order("nome")
        .execute()
        .data
        or []
    )

    for i, product in enumerate(rows, start=2):
        nome = product.get("nome") or ""
        codigo = (product.get("codigo") or "").strip() or make_code(nome, i - 1)
        preco = float(product.get("preco_tabela") or 0)
        preco_min = product.get("preco_minimo")
        estoque = product.get("saldo_estoque")
        if estoque is None:
            estoque = 10
        unidade = product.get("unidade") or "Unidade"
        desc = (product.get("descricao") or "").replace("\n", " ").strip()
        if len(desc) > 500:
            desc = desc[:497] + "..."
        ncm = pick(NCM_MAP, nome, "99999999")
        categoria = pick(CATEGORY_MAP, nome, "Geral")
        ativo = 0 if product.get("ativo", True) else 1

        ws.cell(i, 1, codigo)
        ws.cell(i, 2, nome)
        ws.cell(i, 3, preco)
        if preco_min is not None:
            ws.cell(i, 4, float(preco_min))
        ws.cell(i, 6, ncm)
        if desc:
            ws.cell(i, 8, desc)
        ws.cell(i, 9, unidade)
        ws.cell(i, 10, int(estoque))
        ws.cell(i, 17, categoria)
        ws.cell(i, 20, ativo)

    wb.save(out)
    print(f"Arquivo salvo: {out}")
    print(f"Produtos preenchidos: {len(rows)}")
    for idx, product in enumerate(rows, 1):
        nome = product.get("nome") or ""
        print(f"  {idx}. {make_code(nome, idx)} | {nome[:55]} | NCM {pick(NCM_MAP, nome, '')}")


if __name__ == "__main__":
    main()
